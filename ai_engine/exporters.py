"""
参数导出模块

功能：
1. 多格式导出（JSON, CSV, Excel）
2. 数据集成 API
3. 实时推送接口
"""

import json
import csv
from io import StringIO, BytesIO
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
from abc import ABC, abstractmethod


class Exporter(ABC):
    """导出器基类。"""
    
    @abstractmethod
    def export(self, data: Dict[str, Any]) -> Union[str, bytes]:
        """导出数据。"""
        pass
    
    @abstractmethod
    def get_content_type(self) -> str:
        """获取 MIME 类型。"""
        pass
    
    @abstractmethod
    def get_file_extension(self) -> str:
        """获取文件扩展名。"""
        pass


class JSONExporter(Exporter):
    """JSON 导出器。"""
    
    def export(self, data: Dict[str, Any]) -> str:
        """导出为 JSON。"""
        return json.dumps(data, indent=2, default=str, ensure_ascii=False)
    
    def get_content_type(self) -> str:
        return "application/json"
    
    def get_file_extension(self) -> str:
        return "json"


class CSVExporter(Exporter):
    """CSV 导出器。"""
    
    def export(self, data: Dict[str, Any]) -> str:
        """导出为 CSV。"""
        parameters = data.get("parameters", [])
        
        if not parameters:
            return ""
        
        # 获取所有字段
        fieldnames = set()
        for param in parameters:
            fieldnames.update(param.keys())
        fieldnames = sorted(fieldnames)
        
        # 构建 CSV
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for param in parameters:
            writer.writerow({k: param.get(k, "") for k in fieldnames})
        
        return output.getvalue()
    
    def get_content_type(self) -> str:
        return "text/csv"
    
    def get_file_extension(self) -> str:
        return "csv"


class ExcelExporter(Exporter):
    """Excel 导出器（需要 openpyxl）。"""
    
    def export(self, data: Dict[str, Any]) -> bytes:
        """导出为 Excel。"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "参数"
        
        parameters = data.get("parameters", [])
        
        if not parameters:
            return BytesIO().getvalue()
        
        # 获取所有字段
        fieldnames = set()
        for param in parameters:
            fieldnames.update(param.keys())
        fieldnames = sorted(fieldnames)
        
        # 写表头
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写数据
        for row_idx, param in enumerate(parameters, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                value = param.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 调整列宽
        for col_idx, field in enumerate(fieldnames, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = max(15, len(str(field)))
        
        # 保存到 BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def get_content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    def get_file_extension(self) -> str:
        return "xlsx"


class ExportManager:
    """导出管理器。"""
    
    def __init__(self):
        self.exporters = {
            "json": JSONExporter(),
            "csv": CSVExporter(),
            "xlsx": ExcelExporter(),
        }
    
    def export(
        self,
        data: Dict[str, Any],
        format_type: str = "json",
        debug: bool = False
    ) -> Union[str, bytes]:
        """
        导出数据。
        
        Args:
            data: 待导出的数据
            format_type: 格式类型 ("json" | "csv" | "xlsx")
            debug: 调试模式
        
        Returns:
            导出后的数据
        """
        if format_type not in self.exporters:
            raise ValueError(f"不支持的格式: {format_type}")
        
        exporter = self.exporters[format_type]
        
        try:
            result = exporter.export(data)
            if debug:
                print(f"✅ 导出成功: {format_type.upper()}")
            return result
        except Exception as e:
            if debug:
                print(f"❌ 导出失败: {e}")
            raise
    
    def get_supported_formats(self) -> List[Dict[str, str]]:
        """获取支持的导出格式列表。"""
        formats = []
        for fmt, exporter in self.exporters.items():
            formats.append({
                "format": fmt,
                "extension": exporter.get_file_extension(),
                "content_type": exporter.get_content_type(),
            })
        return formats


class DataIntegrationAPI:
    """数据集成 API。"""
    
    def __init__(self):
        self.export_manager = ExportManager()
        self.webhooks = []  # Webhook 端点列表
    
    def register_webhook(self, url: str, event_types: List[str]) -> None:
        """
        注册 Webhook 回调。
        
        Args:
            url: 回调 URL
            event_types: 事件类型列表 ("parameter_extracted", "anomaly_detected", etc.)
        """
        self.webhooks.append({
            "url": url,
            "event_types": event_types,
            "registered_at": datetime.now().isoformat(),
        })
    
    def trigger_webhook(self, event_type: str, payload: Dict[str, Any]) -> None:
        """
        触发 Webhook 回调。
        
        Args:
            event_type: 事件类型
            payload: 事件负载
        """
        relevant_webhooks = [
            w for w in self.webhooks if event_type in w["event_types"]
        ]
        
        for webhook in relevant_webhooks:
            try:
                import requests
                requests.post(
                    webhook["url"],
                    json={
                        "event_type": event_type,
                        "timestamp": datetime.now().isoformat(),
                        "payload": payload,
                    },
                    timeout=5
                )
            except Exception as e:
                print(f"❌ Webhook 触发失败: {webhook['url']} - {e}")
    
    def create_export_response(
        self,
        data: Dict[str, Any],
        format_type: str = "json"
    ) -> Dict[str, Any]:
        """
        创建导出响应。
        
        Returns:
            {
                "format": 格式类型,
                "extension": 文件扩展名,
                "content_type": MIME 类型,
                "data": 导出数据,
                "size_bytes": 数据大小,
                "timestamp": 导出时间,
            }
        """
        exported = self.export_manager.export(data, format_type)
        
        if isinstance(exported, bytes):
            size = len(exported)
        else:
            size = len(exported.encode("utf-8"))
        
        exporter = self.export_manager.exporters[format_type]
        
        return {
            "format": format_type,
            "extension": exporter.get_file_extension(),
            "content_type": exporter.get_content_type(),
            "data": exported,
            "size_bytes": size,
            "timestamp": datetime.now().isoformat(),
        }
    
    def list_webhooks(self) -> List[Dict[str, Any]]:
        """列出已注册的 Webhook。"""
        return self.webhooks
    
    def clear_webhooks(self) -> None:
        """清除所有 Webhook。"""
        self.webhooks.clear()


# 全局导出管理器实例
_export_manager = ExportManager()
_data_integration_api = DataIntegrationAPI()


def get_export_manager() -> ExportManager:
    """获取全局导出管理器。"""
    global _export_manager
    if _export_manager is None:
        _export_manager = ExportManager()
    return _export_manager


def get_data_integration_api() -> DataIntegrationAPI:
    """获取全局数据集成 API。"""
    global _data_integration_api
    if _data_integration_api is None:
        _data_integration_api = DataIntegrationAPI()
    return _data_integration_api
